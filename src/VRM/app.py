from flask import Flask, Response, render_template, request
from io import BytesIO
import datetime
import plotly.graph_objs as go
from plotly.offline import plot
from dotenv import load_dotenv
from openpyxl import Workbook

import main as vrm_main

load_dotenv()

app = Flask(__name__)

def fill_missing_intervals(series, interval_enum=vrm_main.Interval.MINS15):
    """
    Fills missing intervals in a time series with zero values based on the specified interval enumeration.
    """
    if not isinstance(series, list) or not series:
        return []
    interval_ms = interval_enum.value["ms"]  
    series_sorted = sorted(series, key=lambda x: x[0])
    filled = []
    current = series_sorted[0][0]
    end = series_sorted[-1][0]
    idx = 0
    while current <= end:
        if idx < len(series_sorted) and series_sorted[idx][0] == current:
            filled.append(series_sorted[idx])
            idx += 1
        else:
            filled.append([current, 0])
        current += interval_ms
    return filled


def get_live_records(startdate, enddate, interval):
    """
    Gets live records from the VRM API for the specified date range and interval, and returns the KWH, EVCS, and battery stats records as JSON objects.
    """
    loginobj = vrm_main.login()
    installs = vrm_main.get_installs(loginobj)
    if not installs or 'records' not in installs or not installs['records']:
        raise RuntimeError('No installations found in VRM API response')
    site_id = installs['records'][0]['idSite']

    kwh = vrm_main.get_install_data(loginobj, site_id, startdate, vrm_main.InstallDataType.KWH, interval, enddate)
    evcs = vrm_main.get_install_data(loginobj, site_id, startdate, vrm_main.InstallDataType.EVCS, interval, enddate)
    battery_stats = vrm_main.get_install_data(loginobj, site_id, startdate, vrm_main.InstallDataType.LIVE_FEED_OTHER, interval, enddate)

    kwh_records = kwh.get('records', {}) if isinstance(kwh, dict) else {}
    evcs_records = evcs.get('records', {}) if isinstance(evcs, dict) else {}
    battery_stats_records = battery_stats.get('records', {}) if isinstance(battery_stats, dict) else {}

    # Debug any VRM series with unexpected types (e.g. bool) so we can inspect API data shape.
    for key in ['Gc', 'Pc', 'Bc', 'Pb', 'Gb', 'kwh', 'evE', 'bs']:
        target = None
        if key in kwh_records:
            target = kwh_records[key]
        elif key in evcs_records:
            target = evcs_records[key]
        elif key in battery_stats_records:
            target = battery_stats_records[key]

        if target is not None and not isinstance(target, list):
            print(f"[VRM DEBUG] Unexpected series type for {key}: {type(target).__name__}, value={target}")
            target = []  # Treat as empty series to avoid breaking the app

    for key in ['Gc', 'Pc', 'Bc']:
        if key in kwh_records:
            kwh_records[key] = fill_missing_intervals(kwh_records[key], interval)
    if 'evE' in evcs_records:
        evcs_records['evE'] = fill_missing_intervals(evcs_records['evE'], interval)
    if 'bs' in battery_stats_records:
        battery_stats_records['bs'] = fill_missing_intervals(battery_stats_records['bs'], interval)

    return kwh_records, evcs_records, battery_stats_records


def build_lookup(series):
    """
    Builds a lookup dictionary from a series of [timestamp, value] pairs for easy access to values by timestamp.
    """
    if not isinstance(series, list):
        return {}
    out = {}
    for pt in series:
        if isinstance(pt, (list, tuple)) and len(pt) >= 2:
            key, val = pt[0], pt[1]
            out[key] = val
    return out


def parse_dashboard_params(args):
    """Parses dashboard request parameters into a consistent date/interval/cost set."""
    start_str = args.get('startdate')
    if start_str:
        try:
            startdate = datetime.datetime.strptime(start_str, '%Y-%m-%d').date()
        except ValueError:
            startdate = datetime.date.today() - datetime.timedelta(days=2)
    else:
        startdate = datetime.date.today() - datetime.timedelta(days=2)

    end_str = args.get('enddate')
    if end_str:
        try:
            enddate = datetime.datetime.strptime(end_str + " 23:59:59", '%Y-%m-%d %H:%M:%S')
        except ValueError:
            enddate = datetime.datetime.now()
    else:
        enddate = datetime.datetime.now()

    interval_raw = args.get('interval', '15mins')
    interval_options = {
        '15mins': vrm_main.Interval.MINS15,
        'hours': vrm_main.Interval.HOURS,
        '2hours': vrm_main.Interval.HOURS2,
        'days': vrm_main.Interval.DAYS,
        'weeks': vrm_main.Interval.WEEKS,
        'months': vrm_main.Interval.MONTHS,
        'years': vrm_main.Interval.YEARS,
    }
    interval = interval_options.get(interval_raw, vrm_main.Interval.MINS15)

    cost_per_kwh = 0.3868
    cost_str = args.get('cost_per_kwh')
    if cost_str:
        try:
            cost_per_kwh = float(cost_str)
        except ValueError:
            cost_per_kwh = 0.3868

    return startdate, enddate, interval, interval_raw, cost_per_kwh


def build_dashboard_payload(startdate, enddate, interval, interval_raw, cost_per_kwh):
    """Builds the dashboard payload, including charts and totals, from the live VRM data."""
    errors = []
    status_message = 'Using cached data.'
    battery_stats_records = {}

    try:
        kwh_records, evcs_records, battery_stats_records = get_live_records(startdate, enddate, interval)
        status_message = f'Live data loaded from API (start {startdate} end {enddate} interval {interval}).'
    except Exception as e:
        status_message = f'Live API failed: {e}.'
        errors.append((datetime.datetime.now(), f'Live data failed: {e}', 'critical'))

    if not kwh_records:
        status_message = 'No live data available.'

    bc_lookup = build_lookup(kwh_records.get('Bc', []))
    pc_lookup = build_lookup(kwh_records.get('Pc', []))
    gc_lookup = build_lookup(kwh_records.get('Gc', []))
    pb_lookup = build_lookup(kwh_records.get('Pb', []))
    gb_lookup = build_lookup(kwh_records.get('Gb', []))
    kwh_lookup = build_lookup(kwh_records.get('kwh', []))
    ev_lookup = build_lookup(evcs_records.get('evE', []))
    bs_lookup = build_lookup(battery_stats_records.get('bs', []))

    timestamps = sorted(set(bc_lookup) | set(pc_lookup) | set(gc_lookup) | set(ev_lookup) | set(kwh_lookup) | set(pb_lookup) | set(gb_lookup) | set(bs_lookup))

    times = []
    raw_car = []
    house_no_ev = []
    house_raw = []
    adj_bc = []
    adj_pc = []
    adj_gc = []
    pb_flow = []
    gb_flow = []
    bs_pct = []

    total_car = 0.0
    total_house = 0.0
    total_raw = 0.0
    total_battery = 0.0
    total_solar = 0.0
    total_grid = 0.0

    severity_map = {'critical': 0, 'high': 0, 'medium': 0}
    for ts in timestamps:
        t = datetime.datetime.fromtimestamp(ts / 1000)
        times.append(t)

        car = ev_lookup.get(ts, 0.0)
        bc = bc_lookup.get(ts, 0.0)
        pc = pc_lookup.get(ts, 0.0)
        gc = gc_lookup.get(ts, 0.0)

        pc_adj = pc - car
        bc_adj = bc
        gc_adj = gc

        if pc_adj < 0:
            deficit = -pc_adj
            pc_adj = 0
            bc_adj -= deficit
            if bc_adj < 0:
                deficit = -bc_adj
                bc_adj = 0
                gc_adj -= deficit
                if gc_adj < 0:
                    severity = 'critical'
                    errors.append((t, 'EV removal drives totals negative', severity))
                    severity_map[severity] += 1
                    gc_adj = 0
                else:
                    severity = 'high'
                    errors.append((t, 'Adjusted Gc may be negative after EV removal', severity))
                    severity_map[severity] += 1
            else:
                severity = 'medium'
                errors.append((t, 'Adjusted Bc may be negative after EV removal', severity))
                severity_map[severity] += 1

        if bc_adj < 0 or pc_adj < 0 or gc_adj < 0:
            severity = 'high'
            errors.append((t, 'Adjusted component negative', severity))
            severity_map[severity] += 1

        rest_house = bc_adj + pc_adj + gc_adj
        raw = bc + pc + gc

        raw_car.append(car)
        house_no_ev.append(rest_house)
        house_raw.append(raw)
        adj_bc.append(bc_adj)
        adj_pc.append(pc_adj)
        adj_gc.append(gc_adj)
        pb_flow.append(pb_lookup.get(ts, 0.0))
        gb_flow.append(gb_lookup.get(ts, 0.0))
        bs_pct.append(bs_lookup.get(ts, None))

        total_car += car
        total_house += rest_house
        total_raw += raw
        total_battery += bc_adj
        total_solar += pc_adj
        total_grid += gc_adj

    battery_soc_latest = None
    battery_soc_min = None
    battery_soc_max = None
    consumed_pct_ev = None
    consumed_pct_house = None
    component_pct_battery = None
    component_pct_solar = None
    component_pct_grid = None

    if bs_pct:
        bs_clean = [v for v in bs_pct if isinstance(v, (int, float))]
        if bs_clean:
            battery_soc_latest = bs_clean[-1]
            battery_soc_min = min(bs_clean)
            battery_soc_max = max(bs_clean)

    total_consumed = total_car + total_house
    if total_consumed > 0:
        consumed_pct_ev = (total_car / total_consumed) * 100
        consumed_pct_house = (total_house / total_consumed) * 100

    if total_house > 0:
        component_pct_battery = (total_battery / total_house) * 100
        component_pct_solar = (total_solar / total_house) * 100
        component_pct_grid = (total_grid / total_house) * 100

    ev_cost = total_car * cost_per_kwh
    battery_cost = total_battery * cost_per_kwh
    solar_cost = total_solar * cost_per_kwh
    grid_cost = total_grid * cost_per_kwh
    realized_savings = battery_cost + solar_cost
    estimated_total_savings = realized_savings + ev_cost
    house_cost = total_house * cost_per_kwh
    raw_cost = total_raw * cost_per_kwh

    divs = []
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=times, y=raw_car, mode='lines', name='EV Charge'))
    fig.add_trace(go.Scatter(x=times, y=house_no_ev, mode='lines', name='House (no EV)'))
    fig.add_trace(go.Scatter(x=times, y=house_raw, mode='lines', name='House raw (Bc+Pc+Gc)'))
    fig.update_layout(template='plotly_dark', title='EV vs House Consumption', xaxis=dict(title='Time'), yaxis=dict(title='kWh'), legend=dict(orientation='h', y=-0.2))
    divs.append(plot(fig, output_type='div', include_plotlyjs=False))

    fig2 = go.Figure()
    if times and adj_pc and adj_bc and adj_gc:
        z_data = [adj_pc, adj_bc, adj_gc]
        y_sources = ['Solar', 'Battery', 'Grid']
        fig2.add_trace(go.Surface(x=times, y=y_sources, z=z_data, colorscale='Viridis'))

    fig2.update_layout(
        template='plotly_dark',
        title='House consumption (No EV) 3D source mix',
        scene=dict(
            xaxis=dict(title='Time'),
            yaxis=dict(title='Source'),
            zaxis=dict(title='kWh')
        ),
        margin=dict(l=0, r=0, b=0, t=40)
    )
    divs.append(plot(fig2, output_type='div', include_plotlyjs=False))

    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=times, y=pb_flow, mode='lines', name='Solar'))
    fig3.add_trace(go.Scatter(x=times, y=gb_flow, mode='lines', name='Grid'))
    fig3.update_layout(template='plotly_dark', title='Battery Charging sources', xaxis=dict(title='Time'), yaxis=dict(title='kWh'), legend=dict(orientation='h', y=-0.2))
    divs.append(plot(fig3, output_type='div', include_plotlyjs=False))

    fig4 = go.Figure()
    fig4.add_trace(go.Scatter(x=times, y=bs_pct, mode='lines', name='Battery SOC %', line=dict(color='cyan')))
    fig4.update_layout(template='plotly_dark', title='Battery State of Charge (%)', xaxis=dict(title='Time'), yaxis=dict(title='Battery %', range=[0, 100]), legend=dict(orientation='h', y=-0.2))
    divs.append(plot(fig4, output_type='div', include_plotlyjs=False))

    consumed_values = [total_car, total_house]
    consumed_labels = ['EV', 'House']
    pie1 = go.Figure(go.Pie(labels=consumed_labels, values=consumed_values, hole=0.35, marker=dict(colors=['#82c4ff', '#95e05d'])))
    pie1.update_layout(template='plotly_dark', title='Consumed Split (EV vs House)', margin=dict(t=35, b=5, l=5, r=5), legend=dict(orientation='h', y=-0.2), width=320, height=260)

    house_values = [total_battery, total_solar, total_grid]
    house_labels = ['Battery', 'Solar', 'Grid']
    pie2 = go.Figure(go.Pie(labels=house_labels, values=house_values, hole=0.35, marker=dict(colors=['#00d8ff', '#f8a25d', '#ff8e5c'])))
    pie2.update_layout(template='plotly_dark', title='House Mix', margin=dict(t=35, b=5, l=5, r=5), legend=dict(orientation='h', y=-0.2), width=320, height=260)

    overall_values = [total_car, total_grid, total_solar, total_battery]
    overall_labels = ['EV', 'Grid', 'Solar', 'Battery']
    pie3 = go.Figure(go.Pie(labels=overall_labels, values=overall_values, hole=0.35, marker=dict(colors=['#82c4ff', '#ff8e5c', '#f8a25d', '#00d8ff'])))
    pie3.update_layout(template='plotly_dark', title='EV + Grid + Solar + Battery (%)', margin=dict(t=35, b=5, l=5, r=5), legend=dict(orientation='h', y=-0.2), width=320, height=260)

    pie_graphs = [
        plot(pie1, output_type='div', include_plotlyjs=False),
        plot(pie2, output_type='div', include_plotlyjs=False),
        plot(pie3, output_type='div', include_plotlyjs=False),
    ]

    error_rows = [
        {'time': t.strftime('%Y-%m-%d %H:%M'), 'message': msg, 'severity': severity}
        for t, msg, severity in errors
    ]

    return {
        'graphs': divs,
        'pie_graphs': pie_graphs,
        'totals': {
            'ev_total': total_car,
            'house_total_adjusted': total_house,
            'house_total_raw': total_raw,
            'battery_soc_latest': battery_soc_latest,
            'battery_soc_min': battery_soc_min,
            'battery_soc_max': battery_soc_max,
            'consumed_pct_ev': consumed_pct_ev,
            'consumed_pct_house': consumed_pct_house,
            'component_pct_battery': component_pct_battery,
            'component_pct_solar': component_pct_solar,
            'component_pct_grid': component_pct_grid,
            'cost_per_kwh': cost_per_kwh,
            'ev_cost': ev_cost,
            'battery_cost': battery_cost,
            'solar_cost': solar_cost,
            'grid_cost': grid_cost,
            'battery_kwh': total_battery,
            'solar_kwh': total_solar,
            'grid_kwh': total_grid,
            'realized_savings': realized_savings,
            'estimated_total_savings': estimated_total_savings,
            'house_cost': house_cost,
            'raw_cost': raw_cost,
        },
        'error_rows': error_rows,
        'startdate': startdate.strftime('%Y-%m-%d'),
        'enddate': enddate.strftime('%Y-%m-%d'),
        'interval': interval_raw,
        'severity_map': severity_map,
        'status_message': status_message,
        'kwh_records': kwh_records,
        'evcs_records': evcs_records,
        'battery_stats_records': battery_stats_records,
    }


@app.route('/export')
def export_data():
    """Exports the currently loaded VRM data into an Excel workbook with multiple worksheets."""
    startdate, enddate, interval, interval_raw, cost_per_kwh = parse_dashboard_params(request.args)
    payload = build_dashboard_payload(startdate, enddate, interval, interval_raw, cost_per_kwh)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    summary_sheet.append(['metric', 'value'])
    summary_sheet.append(['status', payload['status_message']])
    summary_sheet.append(['start_date', payload['startdate']])
    summary_sheet.append(['end_date', payload['enddate']])
    summary_sheet.append(['interval', payload['interval']])
    summary_sheet.append(['cost_per_kwh', payload['totals']['cost_per_kwh']])
    for key, value in payload['totals'].items():
        if key != 'cost_per_kwh':
            summary_sheet.append([key, value])
    summary_sheet.append(['severity_critical', payload['severity_map'].get('critical', 0)])
    summary_sheet.append(['severity_high', payload['severity_map'].get('high', 0)])
    summary_sheet.append(['severity_medium', payload['severity_map'].get('medium', 0)])

    def add_series_sheet(name, records):
        sheet = workbook.create_sheet(title=name)
        sheet.append(['timestamp_ms', 'timestamp', 'series', 'value'])
        for series_name, series_values in records.items():
            if not isinstance(series_values, list):
                continue
            for point in series_values:
                if isinstance(point, (list, tuple)) and len(point) >= 2:
                    ts, value = point[0], point[1]
                    timestamp = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
                    sheet.append([ts, timestamp, series_name, value])

    def add_pivoted_sheet(name, records, display_names=None):
        sheet = workbook.create_sheet(title=name)
        display_names = display_names or {}
        rows = {}
        for series_name, series_values in records.items():
            if not isinstance(series_values, list):
                continue
            for point in series_values:
                if isinstance(point, (list, tuple)) and len(point) >= 2:
                    ts, value = point[0], point[1]
                    rows.setdefault(ts, {})[series_name] = value

        headers = ['timestamp_ms', 'timestamp'] + [display_names.get(series_name, series_name) for series_name in records.keys() if isinstance(records.get(series_name), list)]
        sheet.append(headers)
        for ts in sorted(rows):
            row = [ts, datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')]
            for series_name in records.keys():
                if isinstance(records.get(series_name), list):
                    row.append(rows[ts].get(series_name))
            sheet.append(row)

    def add_pivoted_kwh_sheet(records):
        sheet = workbook.create_sheet(title='Raw KWH')
        display_names = {
            'Gc': 'Grid',
            'Pc': 'Solar',
            'Bc': 'Battery',
            'Pb': 'Battery Flow',
            'Gb': 'Grid Flow',
            'kwh': 'kWh',
        }
        rows = {}
        for series_name, series_values in records.items():
            if not isinstance(series_values, list):
                continue
            for point in series_values:
                if isinstance(point, (list, tuple)) and len(point) >= 2:
                    ts, value = point[0], point[1]
                    rows.setdefault(ts, {})[series_name] = value

        headers = ['timestamp_ms', 'timestamp'] + [display_names.get(series_name, series_name) for series_name in records.keys() if isinstance(records.get(series_name), list)]
        sheet.append(headers)
        for ts in sorted(rows):
            row = [ts, datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')]
            for series_name in records.keys():
                if isinstance(records.get(series_name), list):
                    row.append(rows[ts].get(series_name))
            sheet.append(row)

    add_pivoted_kwh_sheet(payload['kwh_records'])
    add_series_sheet('Raw EVCS', payload['evcs_records'])
    add_pivoted_sheet('Raw Battery Stats', payload['battery_stats_records'], {'bs': 'Battery SOC'})

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"vrm-export-{payload['startdate']}-to-{payload['enddate']}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@app.route('/')
def index():
    """
    Renders the main dashboard page with graphs and totals based on live data from the VRM API, using optional query parameters for date range, interval, and cost per kWh.
        Query parameters:
            - startdate: Start date for data retrieval in YYYY-MM-DD format (default: 2 days ago)
            - enddate: End date for data retrieval in YYYY-MM-DD format (default: now)
            - interval: Data interval for aggregation (options: 15mins, hours, 2hours, days, weeks, months, years; default: 15mins)
            - cost_per_kwh: Cost per kWh for calculating costs and savings (default: 0.3868)
    """
    startdate, enddate, interval, interval_raw, cost_per_kwh = parse_dashboard_params(request.args)
    payload = build_dashboard_payload(startdate, enddate, interval, interval_raw, cost_per_kwh)

    return render_template(
        'index.html',
        graphs=payload['graphs'],
        pie_graphs=payload['pie_graphs'],
        totals=payload['totals'],
        error_rows=payload['error_rows'],
        startdate=payload['startdate'],
        enddate=payload['enddate'],
        interval=payload['interval'],
        severity_map=payload['severity_map'],
        status_message=payload['status_message'],
    )


if __name__ == '__main__':
    app.run(debug=True)
