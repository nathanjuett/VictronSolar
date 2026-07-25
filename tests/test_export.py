import os
import sys
import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src", "VRM"))

import app as vrm_app


class ExportRouteTest(unittest.TestCase):
    def setUp(self):
        self.client = vrm_app.app.test_client()

    def test_export_returns_excel_workbook(self):
        sample_kwh = {
            "Gc": [[1_700_000_000_000, 2.0]],
            "Pc": [[1_700_000_000_000, 4.0]],
            "Bc": [[1_700_000_000_000, 1.0]],
            "Pb": [[1_700_000_000_000, 0.0]],
            "Gb": [[1_700_000_000_000, 0.0]],
            "kwh": [[1_700_000_000_000, 5.0]],
        }
        sample_evcs = {"evE": [[1_700_000_000_000, 3.0]]}
        sample_battery = {"bs": [[1_700_000_000_000, 80.0]]}

        with patch.object(vrm_app, "get_live_records", return_value=(sample_kwh, sample_evcs, sample_battery)):
            response = self.client.get("/export?startdate=2025-01-01&enddate=2025-01-02&interval=days&cost_per_kwh=0.5")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.headers["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.data))
        expected_sheets = {
            "Summary",
            "Raw KWH",
            "Raw EVCS",
            "Raw Battery Stats",
            "Raw Venus",
            "Raw Live Feed",
            "Raw Consumption",
            "Raw Solar Yield",
            "Raw Generator",
            "Raw Generator Runtime",
            "Raw Forecast",
        }
        self.assertEqual(set(workbook.sheetnames), expected_sheets)
        raw_kwh_sheet = workbook["Raw KWH"]
        self.assertEqual(raw_kwh_sheet.cell(row=1, column=1).value, "timestamp_ms")
        self.assertEqual(raw_kwh_sheet.cell(row=1, column=3).value, "Grid")
        self.assertEqual(raw_kwh_sheet.cell(row=1, column=4).value, "Solar")

        raw_battery_sheet = workbook["Raw Battery Stats"]
        self.assertEqual(raw_battery_sheet.cell(row=1, column=1).value, "timestamp_ms")
        self.assertEqual(raw_battery_sheet.cell(row=1, column=3).value, "Battery SOC")

    def test_export_flattens_multi_value_series_into_separate_columns(self):
        sample_kwh = {"kwh": [[1_700_000_000_000, 5.0]]}
        sample_evcs = {}
        sample_battery = {}
        multi_value_records = {
            vrm_app.vrm_main.InstallDataType.VENUS: {
                "bs": [[1_700_000_000_000, 77.23529411764706, 74.0, 80.0]],
                "consumption": [[1_700_000_000_000, 10.0, 12.5]],
            }
        }

        with patch.object(vrm_app, "get_live_records", return_value=(sample_kwh, sample_evcs, sample_battery)):
            with patch.object(vrm_app, "get_export_records", return_value=multi_value_records):
                response = self.client.get("/export?startdate=2025-01-01&enddate=2025-01-02&interval=days&cost_per_kwh=0.5")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        venus_sheet = workbook["Raw Venus"]
        self.assertEqual(venus_sheet.cell(row=1, column=3).value, "bs_0")
        self.assertEqual(venus_sheet.cell(row=1, column=4).value, "bs_1")
        self.assertEqual(venus_sheet.cell(row=1, column=5).value, "bs_2")
        self.assertEqual(venus_sheet.cell(row=1, column=6).value, "consumption_0")
        self.assertEqual(venus_sheet.cell(row=1, column=7).value, "consumption_1")
        self.assertEqual(venus_sheet.cell(row=2, column=3).value, 77.23529411764706)
        self.assertEqual(venus_sheet.cell(row=2, column=4).value, 74.0)
        self.assertEqual(venus_sheet.cell(row=2, column=5).value, 80.0)
        self.assertEqual(venus_sheet.cell(row=2, column=6).value, 10.0)
        self.assertEqual(venus_sheet.cell(row=2, column=7).value, 12.5)


if __name__ == "__main__":
    unittest.main()
